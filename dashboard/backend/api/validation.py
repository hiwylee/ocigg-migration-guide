from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone
import aiosqlite
import os

from core.db import get_db
from api.auth import get_current_user, UserInfo

router = APIRouter()

# Whitelist of columns that may be updated via PATCH /items/{id}
_ALLOWED_UPDATE_COLS = {"status", "note", "assignee", "verified_at", "verified_by"}

# ---------------------------------------------------------------------------
# Domain definitions
# ---------------------------------------------------------------------------

DOMAINS = [
    ("01_GG_Process",       28),
    ("02_Static_Schema",    38),
    ("03_Data_Validation",  25),
    ("04_Special_Objects",  42),
    ("05_Migration_Caution",26),
]

DOMAIN_ITEMS = {
    "01_GG_Process": [
        ("GoldenGate Extract 프로세스 기동 전 점검",        "HIGH",   "GG 명령: INFO EXTRACT *"),
        ("Extract 파라미터 파일 검증",                     "HIGH",   "파라미터 파일 확인"),
        ("Extract SCN 기준점 확인",                        "HIGH",   "GG 명령: INFO EXTRACT, SHOWCH"),
        ("Supplemental Logging 활성화 확인 (MIN)",         "HIGH",   "SQL: V$DATABASE 조회"),
        ("PK 없는 테이블 ALL COLUMNS Supplemental Logging","HIGH",   "SQL: DBA_LOG_GROUPS 조회"),
        ("Streams Pool 크기 확인 (≥256MB)",                "HIGH",   "SQL: V$SGA_DYNAMIC_COMPONENTS 조회"),
        ("ENABLE_GOLDENGATE_REPLICATION 파라미터 확인",    "HIGH",   "SQL: V$PARAMETER 조회"),
        ("Extract Trail 파일 생성 확인",                   "MEDIUM", "GG 명령: INFO EXTRACT, SHOWCH"),
        ("Extract 통계 초기 수집",                         "MEDIUM", "GG 명령: STATS EXTRACT"),
        ("Data Pump 프로세스 기동 전 점검",                "HIGH",   "GG 명령: INFO PUMP *"),
        ("Pump 파라미터 파일 검증",                        "HIGH",   "파라미터 파일 확인"),
        ("Pump Trail 전송 확인",                           "HIGH",   "GG 명령: INFO PUMP, SHOWCH"),
        ("네트워크 구간 지연 측정 (AWS→OCI)",              "MEDIUM", "ping / traceroute 실행"),
        ("Replicat 프로세스 기동 전 점검",                 "HIGH",   "GG 명령: INFO REPLICAT *"),
        ("Replicat 파라미터 파일 검증",                    "HIGH",   "파라미터 파일 확인"),
        ("Replicat 적용 모드 확인 (INTEGRATED/CLASSIC)",   "MEDIUM", "GG 명령: INFO REPLICAT, DETAIL"),
        ("Replicat Checkpoint 위치 확인",                  "MEDIUM", "GG 명령: INFO REPLICAT, SHOWCH"),
        ("GG LAG 실시간 모니터링 기준값 설정",             "HIGH",   "GG Admin API 확인"),
        ("Extract LAG < 30초 확인",                        "HIGH",   "GG 명령: LAG EXTRACT"),
        ("Pump LAG < 30초 확인",                           "HIGH",   "GG 명령: LAG PUMP"),
        ("Replicat LAG < 30초 확인",                       "HIGH",   "GG 명령: LAG REPLICAT"),
        ("24시간 LAG 안정화 확인",                         "HIGH",   "대시보드 LAG 이력 차트"),
        ("GoldenGate 에러 로그 점검",                      "MEDIUM", "GG 로그 파일 확인"),
        ("Discard 파일 크기 확인",                         "MEDIUM", "OS 명령: ls -lh *.dsc"),
        ("GG 통계 (Inserts/Updates/Deletes) 정합성 확인",  "HIGH",   "GG 명령: STATS REPLICAT, TOTAL"),
        ("GG 관리 계정 권한 확인 (src/tgt)",               "MEDIUM", "SQL: SESSION_PRIVS 조회"),
        ("GG Deployment 상태 확인",                        "HIGH",   "GG Admin API: /deployments"),
        ("GG 버전 호환성 확인",                            "MEDIUM", "GG Admin API: /version"),
    ],
    "02_Static_Schema": [
        ("테이블 수 비교 (src vs tgt)",                    "HIGH",   "SQL: DBA_TABLES COUNT 비교"),
        ("컬럼 정의 비교 (데이터타입, NOT NULL)",          "HIGH",   "SQL: DBA_TAB_COLUMNS 비교"),
        ("PK 제약조건 비교",                               "HIGH",   "SQL: DBA_CONSTRAINTS TYPE='P' 비교"),
        ("UK 제약조건 비교",                               "HIGH",   "SQL: DBA_CONSTRAINTS TYPE='U' 비교"),
        ("FK 제약조건 비교",                               "HIGH",   "SQL: DBA_CONSTRAINTS TYPE='R' 비교"),
        ("CHECK 제약조건 비교",                            "MEDIUM", "SQL: DBA_CONSTRAINTS TYPE='C' 비교"),
        ("인덱스 목록 비교 (개수, 이름)",                  "HIGH",   "SQL: DBA_INDEXES 비교"),
        ("인덱스 컬럼 구성 비교",                          "HIGH",   "SQL: DBA_IND_COLUMNS 비교"),
        ("BITMAP 인덱스 존재 여부 확인 (SE 미지원)",       "HIGH",   "SQL: DBA_INDEXES WHERE INDEX_TYPE='BITMAP'"),
        ("Function-Based 인덱스 비교",                     "MEDIUM", "SQL: DBA_IND_EXPRESSIONS 비교"),
        ("시퀀스 목록 및 현재값 비교",                     "HIGH",   "SQL: DBA_SEQUENCES 비교"),
        ("시퀀스 INCREMENT BY / NOCYCLE 설정 비교",        "MEDIUM", "SQL: DBA_SEQUENCES 비교"),
        ("뷰(VIEW) 목록 및 정의 비교",                     "MEDIUM", "SQL: DBA_VIEWS 비교"),
        ("뷰 컴파일 오류 확인",                            "HIGH",   "SQL: DBA_OBJECTS WHERE STATUS='INVALID'"),
        ("동의어(SYNONYM) 비교",                           "MEDIUM", "SQL: DBA_SYNONYMS 비교"),
        ("공개 동의어(PUBLIC SYNONYM) 비교",               "LOW",    "SQL: ALL_SYNONYMS WHERE OWNER='PUBLIC'"),
        ("저장 프로시저 목록 비교",                        "HIGH",   "SQL: DBA_PROCEDURES 비교"),
        ("함수(FUNCTION) 목록 비교",                       "HIGH",   "SQL: DBA_OBJECTS TYPE='FUNCTION' 비교"),
        ("패키지(PACKAGE) 목록 비교",                      "HIGH",   "SQL: DBA_OBJECTS TYPE='PACKAGE' 비교"),
        ("패키지 바디(PACKAGE BODY) 컴파일 상태 확인",     "HIGH",   "SQL: DBA_OBJECTS TYPE='PACKAGE BODY' 확인"),
        ("타입(TYPE) 목록 비교",                           "MEDIUM", "SQL: DBA_TYPES 비교"),
        ("트리거(TRIGGER) 활성화 상태 비교",               "HIGH",   "SQL: DBA_TRIGGERS 비교"),
        ("INVALID 객체 수 확인 (tgt ≤ src)",               "HIGH",   "SQL: DBA_OBJECTS STATUS='INVALID' COUNT"),
        ("INVALID 객체 상세 목록 분석",                    "HIGH",   "SQL: DBA_OBJECTS STATUS='INVALID' 목록"),
        ("테이블 스페이스 여유 공간 확인",                 "HIGH",   "SQL: DBA_FREE_SPACE 조회"),
        ("스키마별 세그먼트 크기 비교",                    "MEDIUM", "SQL: DBA_SEGMENTS 비교"),
        ("파티션 테이블 목록 비교",                        "HIGH",   "SQL: DBA_PART_TABLES 비교"),
        ("파티션 타입 비교 (RANGE/LIST/HASH)",             "MEDIUM", "SQL: DBA_PART_TABLES PARTITIONING_TYPE 비교"),
        ("LOB 컬럼 스토리지 파라미터 비교",                "MEDIUM", "SQL: DBA_LOBS 비교"),
        ("클러스터(CLUSTER) 객체 확인",                    "LOW",    "SQL: DBA_CLUSTERS 조회"),
        ("데이터베이스 링크(DB LINK) 목록 비교",           "HIGH",   "SQL: DBA_DB_LINKS 비교"),
        ("디렉터리(DIRECTORY) 객체 비교",                  "MEDIUM", "SQL: DBA_DIRECTORIES 비교"),
        ("스키마 DDL 전체 Export 비교",                    "MEDIUM", "DBMS_METADATA.GET_DDL 실행"),
        ("NLS 파라미터 비교 (src vs tgt)",                 "HIGH",   "SQL: NLS_DATABASE_PARAMETERS 비교"),
        ("문자셋 확인 (AL32UTF8 등)",                      "HIGH",   "SQL: NLS_DATABASE_PARAMETERS NLS_CHARACTERSET"),
        ("권한(GRANT) 비교 (객체 권한)",                   "HIGH",   "SQL: DBA_TAB_PRIVS 비교"),
        ("시스템 권한 비교",                               "MEDIUM", "SQL: DBA_SYS_PRIVS 비교"),
        ("롤(ROLE) 비교",                                  "MEDIUM", "SQL: DBA_ROLES 비교"),
    ],
    "03_Data_Validation": [
        ("전체 테이블 행 수 비교",                         "HIGH",   "SQL: COUNT(*) 전체 비교"),
        ("도메인별 핵심 테이블 행 수 비교",                "HIGH",   "SQL: COUNT(*) 핵심 테이블"),
        ("샘플 체크섬 비교 (상위 N개 테이블)",             "HIGH",   "SQL: ORA_HASH 체크섬 비교"),
        ("LOB 데이터 크기 비교",                           "HIGH",   "SQL: DBMS_LOB.GETLENGTH SUM"),
        ("LOB 데이터 샘플 해시 비교",                      "HIGH",   "SQL: DBMS_CRYPTO.HASH 비교"),
        ("날짜 컬럼 범위 비교 (MIN/MAX)",                  "MEDIUM", "SQL: MIN/MAX DATE 비교"),
        ("숫자 컬럼 합계 비교 (SUM 샘플)",                 "MEDIUM", "SQL: SUM 비교"),
        ("NULL 비율 비교 (주요 컬럼)",                     "MEDIUM", "SQL: NULL COUNT 비교"),
        ("참조 무결성 확인 (FK 위반 행 수)",               "HIGH",   "SQL: FK 위반 행 수 조회"),
        ("중복 PK 확인 (tgt)",                             "HIGH",   "SQL: GROUP BY PK HAVING COUNT > 1"),
        ("최신 삽입/수정 데이터 반영 확인 (Cutover 직전)", "HIGH",   "SQL: MAX 변경 타임스탬프 비교"),
        ("NLS_DATE_FORMAT 변환 확인",                      "MEDIUM", "SQL: TO_CHAR(date) 결과 비교"),
        ("타임존 처리 확인 (TIMESTAMP WITH TZ)",           "HIGH",   "SQL: SYS_EXTRACT_UTC 비교"),
        ("CLOB 한글 데이터 무결성 확인",                   "HIGH",   "SQL: LENGTHB/LENGTH 비교"),
        ("BLOB 바이너리 데이터 무결성 확인",               "HIGH",   "SQL: DBMS_LOB.COMPARE 실행"),
        ("초기 적재(impdp) 완료 행 수 확인",               "HIGH",   "impdp 로그 확인 + SQL COUNT"),
        ("델타 적용 후 행 수 재확인",                      "HIGH",   "SQL: COUNT(*) 재비교"),
        ("GG Replicat 적용 건수 vs Extract 발생 건수 비교","HIGH",   "GG 명령: STATS 비교"),
        ("트랜잭션 일관성 확인 (COMMIT 단위)",             "HIGH",   "GG 트레일 파일 분석"),
        ("대용량 테이블 청크 비교",                        "MEDIUM", "SQL: ROWNUM 범위 COUNT 비교"),
        ("시퀀스 현재값 동기화 확인",                      "HIGH",   "SQL: LAST_NUMBER 비교"),
        ("파티션별 행 수 비교",                            "MEDIUM", "SQL: DBA_TAB_PARTITIONS COUNT 비교"),
        ("삭제된 행 동기화 확인",                          "HIGH",   "GG 통계 + 행 수 차이 분석"),
        ("업데이트된 행 샘플 비교",                        "HIGH",   "SQL: 샘플 행 값 직접 비교"),
        ("배치 처리 결과 데이터 비교",                     "MEDIUM", "SQL: 배치 결과 테이블 비교"),
    ],
    "04_Special_Objects": [
        ("Materialized View 목록 비교",                    "HIGH",   "SQL: DBA_MVIEWS 비교"),
        ("MV 새로고침 모드 확인 (FAST/COMPLETE/FORCE)",    "HIGH",   "SQL: DBA_MVIEWS.REFRESH_METHOD 비교"),
        ("MV 새로고침 그룹 확인",                          "MEDIUM", "SQL: DBA_REFRESH 비교"),
        ("MV Log 존재 여부 확인",                          "HIGH",   "SQL: DBA_MVIEW_LOGS 비교"),
        ("MV 데이터 행 수 비교",                           "HIGH",   "SQL: COUNT(*) MV 비교"),
        ("MV 컴파일 오류 확인",                            "HIGH",   "SQL: DBA_OBJECTS STATUS='INVALID'"),
        ("트리거 목록 및 활성화 상태 비교",                "HIGH",   "SQL: DBA_TRIGGERS 비교"),
        ("트리거 실행 순서 확인 (FOLLOWS/PRECEDES)",       "MEDIUM", "SQL: DBA_TRIGGERS.ORDERING 비교"),
        ("트리거 타입 확인 (BEFORE/AFTER/INSTEAD OF)",     "MEDIUM", "SQL: DBA_TRIGGERS.TRIGGER_TYPE 비교"),
        ("DDL 트리거 처리 방안 확인",                      "HIGH",   "SQL: DBA_TRIGGERS WHERE TRIGGERING_EVENT LIKE '%DDL%'"),
        ("DB Link 접속 테스트 (tgt에서)",                  "HIGH",   "SQL: SELECT 1 FROM DUAL@dblink"),
        ("DB Link 대상 호스트 IP 변경 여부 확인",          "HIGH",   "SQL: DBA_DB_LINKS.HOST 확인"),
        ("DB Link를 사용하는 객체 목록 확인",              "HIGH",   "SQL: DBA_DEPENDENCIES REFERENCED_LINK_NAME"),
        ("DBMS_JOB 목록 확인 (소스)",                      "HIGH",   "SQL: DBA_JOBS 조회"),
        ("DBMS_JOB → DBMS_SCHEDULER 변환 계획 확인",       "HIGH",   "문서 검토"),
        ("DBMS_SCHEDULER Job 목록 비교 (tgt)",             "HIGH",   "SQL: DBA_SCHEDULER_JOBS 비교"),
        ("스케줄러 Job 활성화 상태 확인 (tgt)",            "HIGH",   "SQL: DBA_SCHEDULER_JOBS.ENABLED 확인"),
        ("스케줄러 실행 이력 확인 (tgt)",                  "MEDIUM", "SQL: DBA_SCHEDULER_JOB_LOG 확인"),
        ("파티션 테이블 익스텐트 크기 확인",               "MEDIUM", "SQL: DBA_EXTENTS 조회"),
        ("파티션 인덱스 상태 확인 (USABLE/UNUSABLE)",      "HIGH",   "SQL: DBA_IND_PARTITIONS.STATUS 확인"),
        ("파티션 Add/Drop DDL 복제 가능 여부 확인",        "HIGH",   "GG DDL 복제 테스트"),
        ("Range 파티션 MAXVALUE 처리 확인",                "MEDIUM", "SQL: DBA_TAB_PARTITIONS HIGH_VALUE='MAXVALUE'"),
        ("List 파티션 DEFAULT 처리 확인",                  "MEDIUM", "SQL: DBA_TAB_PARTITIONS HIGH_VALUE='DEFAULT'"),
        ("Hash 파티션 수 일치 여부 확인",                  "MEDIUM", "SQL: DBA_TAB_PARTITIONS COUNT 비교"),
        ("Interval 파티션 처리 확인",                      "HIGH",   "SQL: DBA_PART_TABLES WHERE INTERVAL IS NOT NULL"),
        ("GG DDL 복제 범위 확인 (TABLE 객체만)",           "HIGH",   "GG 파라미터 파일 확인"),
        ("GG DDL 필터 설정 확인",                          "HIGH",   "GG 파라미터 파일 확인"),
        ("DDL 히스토리 로그 설정 확인",                    "MEDIUM", "GG 로그 파일 확인"),
        ("External Table 처리 방안 확인",                  "MEDIUM", "SQL: DBA_EXTERNAL_TABLES 조회"),
        ("임시 테이블 (GTT) 처리 방안 확인",               "MEDIUM", "SQL: DBA_TABLES WHERE TEMPORARY='Y'"),
        ("Advanced Queue (AQ) 처리 방안 확인",             "HIGH",   "SQL: DBA_QUEUES 조회"),
        ("Java Stored Procedure 처리 방안 확인",           "MEDIUM", "SQL: DBA_OBJECTS TYPE='JAVA CLASS'"),
        ("Workspace Manager 사용 여부 확인",               "LOW",    "SQL: DBA_WM_VER_TABS 조회"),
        ("Label Security 사용 여부 확인",                  "LOW",    "SQL: DBA_SA_POLICIES 조회"),
        ("Fine-Grained Auditing 설정 비교",                "MEDIUM", "SQL: DBA_AUDIT_POLICIES 비교"),
        ("Audit Trail 이관 계획 확인",                     "LOW",    "문서 검토"),
        ("통계 정보 (DBMS_STATS) 이관 확인",               "HIGH",   "SQL: DBA_TAB_STATISTICS 비교"),
        ("히스토그램 정보 비교",                           "MEDIUM", "SQL: DBA_TAB_COL_STATISTICS 비교"),
        ("Fixed Object 통계 수집 확인",                    "LOW",    "DBMS_STATS.GATHER_FIXED_OBJECTS_STATS 실행"),
        ("AWR/ASH 데이터 이관 계획 확인",                  "LOW",    "문서 검토"),
        ("Recycle Bin 정리 확인",                          "LOW",    "SQL: PURGE RECYCLEBIN 실행"),
        ("Redo Log 크기 및 그룹 수 확인 (tgt)",            "MEDIUM", "SQL: V$LOG 조회"),
    ],
    "05_Migration_Caution": [
        ("Oracle SE 미지원 기능 최종 확인",                "HIGH",   "문서 검토 + SQL 확인"),
        ("BITMAP 인덱스 미사용 확인",                      "HIGH",   "SQL: DBA_INDEXES TYPE='BITMAP'"),
        ("Parallel DML 제한 확인 (SE)",                    "HIGH",   "SQL: V$PARAMETER parallel_max_servers"),
        ("Advanced Compression 미사용 확인 (SE)",          "HIGH",   "SQL: DBA_TABLES WHERE COMPRESS='ENABLED'"),
        ("Partitioning 라이선스 확인 (SE2)",               "HIGH",   "SQL: V$OPTION WHERE PARAMETER='Partitioning'"),
        ("GoldenGate SE 제약 최종 확인",                   "HIGH",   "GG 문서 검토"),
        ("GG Integrated Capture 가용 여부 확인",           "HIGH",   "SQL: V$DATABASE.LOG_MODE + GG 설정"),
        ("DDL 복제 제한 사항 문서화",                      "HIGH",   "문서 검토"),
        ("시퀀스 캐시로 인한 GAP 허용 여부 확인",         "MEDIUM", "업무 담당자 확인"),
        ("UDT (User Defined Type) GG 복제 가능 여부",      "HIGH",   "SQL: DBA_TYPES + GG 테스트"),
        ("XML 타입 컬럼 GG 복제 가능 여부",                "HIGH",   "SQL: DBA_TAB_COLUMNS DATA_TYPE='XMLTYPE'"),
        ("LONG/LONG RAW 컬럼 처리 방안 확인",              "HIGH",   "SQL: DBA_TAB_COLUMNS DATA_TYPE LIKE 'LONG%'"),
        ("BFILE 컬럼 처리 방안 확인",                      "MEDIUM", "SQL: DBA_TAB_COLUMNS DATA_TYPE='BFILE'"),
        ("Deferred Constraint 처리 방안 확인",             "HIGH",   "SQL: DBA_CONSTRAINTS WHERE DEFERRABLE='DEFERRABLE'"),
        ("Supplemental Log 오버헤드 측정",                 "MEDIUM", "SQL: V$SYSSTAT 모니터링"),
        ("소스 DB I/O 증가율 모니터링",                    "MEDIUM", "SQL: V$FILESTAT 모니터링"),
        ("RDS Parameter Group 변경 이력 확인",             "HIGH",   "AWS 콘솔 확인"),
        ("RDS 유지보수 윈도우 충돌 확인",                  "HIGH",   "AWS 콘솔 확인"),
        ("RDS 자동 백업과 expdp 시점 충돌 확인",           "HIGH",   "AWS 콘솔 + 스케줄 확인"),
        ("FastConnect/VPN 대역폭 충분성 확인",             "HIGH",   "네트워크 성능 테스트"),
        ("OCI Security List / NSG 규칙 확인",              "HIGH",   "OCI 콘솔 확인"),
        ("OCI DBCS Backup 설정 확인",                      "HIGH",   "OCI 콘솔 확인"),
        ("롤백 절차 문서화 및 테스트 완료 확인",           "HIGH",   "문서 검토"),
        ("소스 RDS 2주 보존 계획 확인",                    "HIGH",   "AWS 콘솔 + 문서 확인"),
        ("Cut-over 공지 및 다운타임 동의 확인",            "HIGH",   "이해관계자 확인"),
        ("비상 연락망 확인",                               "HIGH",   "문서 검토"),
    ],
}


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class DomainSummary(BaseModel):
    domain: str
    total: int
    pass_count: int
    warn_count: int
    fail_count: int
    pending_count: int


class ValidationSummaryResponse(BaseModel):
    total: int
    pass_count: int
    warn_count: int
    fail_count: int
    pending_count: int
    go_nogo: str
    domain_summary: List[DomainSummary]


class ValidationItemResponse(BaseModel):
    id: int
    domain: str
    item_no: int
    item_name: str
    priority: str
    status: str
    area: Optional[str] = None
    method: Optional[str] = None
    note: Optional[str] = None
    assignee: Optional[str] = None
    verified_at: Optional[str] = None
    verified_by: Optional[str] = None


class ValidationUpdateRequest(BaseModel):
    status: Optional[str] = None
    note: Optional[str] = None
    assignee: Optional[str] = None


class VerifyRequest(BaseModel):
    status: str
    note: Optional[str] = None


# ---------------------------------------------------------------------------
# Go/No-Go logic
# ---------------------------------------------------------------------------

def compute_go_nogo(total: int, pass_count: int, warn_count: int, fail_count: int,
                    pending_count: int, high_all_pass: bool) -> str:
    if fail_count > 0:
        return "NO_GO"
    if pending_count > 0:
        return "PENDING"
    if pass_count == total:
        return "GO"
    if high_all_pass and warn_count > 0:
        return "CONDITIONAL_GO"
    return "PENDING"


# ---------------------------------------------------------------------------
# Seed logic
# ---------------------------------------------------------------------------

async def seed_validation_data() -> None:
    """Seed validation_results table from xlsx or dummy data if empty."""
    from core.db import get_db_path
    import aiosqlite as _aio

    async with _aio.connect(get_db_path()) as db:
        db.row_factory = _aio.Row
        row = await (await db.execute("SELECT COUNT(*) as cnt FROM validation_results")).fetchone()
        if row["cnt"] > 0:
            return

        xlsx_path = "/app/docs/00. validation_plan.xlsx"
        items: List[tuple] = []

        if os.path.exists(xlsx_path):
            try:
                items = _parse_xlsx(xlsx_path)
            except Exception:
                items = []

        if not items:
            items = _generate_dummy_items()

        await db.executemany(
            "INSERT INTO validation_results (domain, item_no, area, item_name, priority, method, status) VALUES (?,?,?,?,?,?,?)",
            items,
        )
        await db.commit()


def _parse_xlsx(path: str) -> List[tuple]:
    """Parse xlsx, return (domain, item_no, area, item_name, priority, method, status)."""
    try:
        import openpyxl
    except ImportError:
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result: List[tuple] = []

    priority_map = {"상": "HIGH", "중": "MEDIUM", "하": "LOW"}

    domain_order = [
        "01_GG_Process", "02_Static_Schema", "03_Data_Validation",
        "04_Special_Objects", "05_Migration_Caution",
    ]

    for sheet_name in wb.sheetnames:
        sn_clean = sheet_name.strip().lstrip("📋").strip()
        if sn_clean not in domain_order:
            continue
        domain = sn_clean

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Detect if col[2] is "단계" (04_Special_Objects structure)
        header = rows[0]
        has_step_col = (
            header and len(header) > 2 and header[2] is not None
            and "단계" in str(header[2])
        )

        item_no = 0
        for row in rows[1:]:
            if not row or all(c is None for c in row):
                continue
            no_val = row[0]
            if no_val is None:
                continue
            if not str(no_val).strip().isdigit():
                continue  # skip section title rows like "▶ ..."

            area = str(row[1]).strip() if row[1] is not None else None

            if has_step_col:
                # 04_Special_Objects: NO | 객체유형 | 단계 | 검증항목 | 중요도
                item_name = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                method = str(row[2]).strip() if len(row) > 2 and row[2] else None
            else:
                # standard: NO | 영역 | 검증항목 | 검증방법 | 중요도
                item_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                method = str(row[3]).strip() if len(row) > 3 and row[3] else None

            if not item_name:
                continue

            priority_raw = str(row[4]).strip() if len(row) > 4 and row[4] is not None else "중"
            priority = priority_map.get(priority_raw, "MEDIUM")

            item_no += 1
            result.append((domain, item_no, area or None, item_name, priority, method or None, "PENDING"))

    return result


def _generate_dummy_items() -> List[tuple]:
    """Generate dummy validation items from DOMAIN_ITEMS definition."""
    result: List[tuple] = []
    for domain, items_list in DOMAIN_ITEMS.items():
        for idx, (name, priority, method) in enumerate(items_list, start=1):
            result.append((domain, idx, None, name, priority, method, "PENDING"))
    return result


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/summary", response_model=ValidationSummaryResponse)
async def get_summary(db: aiosqlite.Connection = Depends(get_db)):
    # Overall counts
    row = await (await db.execute("""
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status='PASS'    THEN 1 ELSE 0 END) as pass_count,
            SUM(CASE WHEN status='WARN'    THEN 1 ELSE 0 END) as warn_count,
            SUM(CASE WHEN status='FAIL'    THEN 1 ELSE 0 END) as fail_count,
            SUM(CASE WHEN status='PENDING' THEN 1 ELSE 0 END) as pending_count
        FROM validation_results
    """)).fetchone()

    total        = row["total"]        or 0
    pass_count   = row["pass_count"]   or 0
    warn_count   = row["warn_count"]   or 0
    fail_count   = row["fail_count"]   or 0
    pending_count= row["pending_count"] or 0

    # Check if all HIGH priority items pass
    high_row = await (await db.execute("""
        SELECT
            COUNT(*) as total_high,
            SUM(CASE WHEN status='PASS' THEN 1 ELSE 0 END) as high_pass
        FROM validation_results
        WHERE priority='HIGH'
    """)).fetchone()
    total_high = high_row["total_high"] or 0
    high_pass  = high_row["high_pass"]  or 0
    high_all_pass = (total_high > 0 and high_pass == total_high)

    go_nogo = compute_go_nogo(total, pass_count, warn_count, fail_count, pending_count, high_all_pass)

    # Domain breakdown
    domain_rows = await (await db.execute("""
        SELECT
            domain,
            COUNT(*) as total,
            SUM(CASE WHEN status='PASS'    THEN 1 ELSE 0 END) as pass_count,
            SUM(CASE WHEN status='WARN'    THEN 1 ELSE 0 END) as warn_count,
            SUM(CASE WHEN status='FAIL'    THEN 1 ELSE 0 END) as fail_count,
            SUM(CASE WHEN status='PENDING' THEN 1 ELSE 0 END) as pending_count
        FROM validation_results
        GROUP BY domain
        ORDER BY domain
    """)).fetchall()

    domain_summary = [
        DomainSummary(
            domain=r["domain"],
            total=r["total"] or 0,
            pass_count=r["pass_count"] or 0,
            warn_count=r["warn_count"] or 0,
            fail_count=r["fail_count"] or 0,
            pending_count=r["pending_count"] or 0,
        )
        for r in domain_rows
    ]

    return ValidationSummaryResponse(
        total=total,
        pass_count=pass_count,
        warn_count=warn_count,
        fail_count=fail_count,
        pending_count=pending_count,
        go_nogo=go_nogo,
        domain_summary=domain_summary,
    )


@router.get("/items", response_model=List[ValidationItemResponse])
async def list_items(
    domain:   Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    status:   Optional[str] = Query(None),
    db: aiosqlite.Connection = Depends(get_db),
):
    query = "SELECT * FROM validation_results WHERE 1=1"
    params: list = []
    if domain:
        query += " AND domain=?"
        params.append(domain)
    if priority:
        query += " AND priority=?"
        params.append(priority)
    if status:
        query += " AND status=?"
        params.append(status)
    query += " ORDER BY domain, item_no"

    rows = await (await db.execute(query, params)).fetchall()
    return [_row_to_item(r) for r in rows]


@router.get("/items/{item_id}", response_model=ValidationItemResponse)
async def get_item(item_id: int, db: aiosqlite.Connection = Depends(get_db)):
    row = await (
        await db.execute("SELECT * FROM validation_results WHERE id=?", (item_id,))
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="항목을 찾을 수 없습니다")
    return _row_to_item(row)


@router.patch("/items/{item_id}", response_model=ValidationItemResponse)
async def update_item(
    item_id: int,
    body: ValidationUpdateRequest,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user),
):
    row = await (
        await db.execute("SELECT * FROM validation_results WHERE id=?", (item_id,))
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="항목을 찾을 수 없습니다")

    now = datetime.now(timezone.utc).isoformat()
    fields: list[str] = []
    params: list = []

    if body.status is not None:
        if body.status not in ("PASS", "WARN", "FAIL", "PENDING"):
            raise HTTPException(status_code=400, detail="status는 PASS|WARN|FAIL|PENDING 중 하나여야 합니다")
        fields += ["status=?", "verified_at=?", "verified_by=?"]
        params += [body.status, now, current_user.username]
    if body.note is not None:
        fields.append("note=?")
        params.append(body.note)
    if body.assignee is not None:
        fields.append("assignee=?")
        params.append(body.assignee)

    if not fields:
        return _row_to_item(row)

    col_names = {f.split("=")[0] for f in fields}
    if not col_names.issubset(_ALLOWED_UPDATE_COLS):
        raise HTTPException(status_code=400, detail="허용되지 않은 필드")

    params.append(item_id)
    await db.execute(f"UPDATE validation_results SET {', '.join(fields)} WHERE id=?", params)

    # event log
    if body.status is not None:
        await db.execute(
            """INSERT INTO event_log (event_type, message, related_item, actor, created_at)
               VALUES ('VALIDATION_UPDATE', ?, ?, ?, ?)""",
            (
                f"[{row['domain']}] {row['item_name']} 상태 변경: {row['status']} → {body.status}",
                str(item_id),
                current_user.username,
                now,
            ),
        )

    await db.commit()

    updated = await (
        await db.execute("SELECT * FROM validation_results WHERE id=?", (item_id,))
    ).fetchone()
    return _row_to_item(updated)


@router.post("/items/{item_id}/verify", response_model=ValidationItemResponse)
async def verify_item(
    item_id: int,
    body: VerifyRequest,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user),
):
    if body.status not in ("PASS", "WARN", "FAIL"):
        raise HTTPException(status_code=400, detail="status는 PASS|WARN|FAIL 중 하나여야 합니다")

    row = await (
        await db.execute("SELECT * FROM validation_results WHERE id=?", (item_id,))
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="항목을 찾을 수 없습니다")

    now = datetime.now(timezone.utc).isoformat()
    await db.execute(
        """UPDATE validation_results
           SET status=?, note=?, verified_at=?, verified_by=?
           WHERE id=?""",
        (body.status, body.note, now, current_user.username, item_id),
    )
    await db.execute(
        """INSERT INTO event_log (event_type, message, related_item, actor, created_at)
           VALUES ('VALIDATION_UPDATE', ?, ?, ?, ?)""",
        (
            f"[{row['domain']}] {row['item_name']} 검증 완료: {body.status}",
            str(item_id),
            current_user.username,
            now,
        ),
    )
    await db.commit()

    updated = await (
        await db.execute("SELECT * FROM validation_results WHERE id=?", (item_id,))
    ).fetchone()
    return _row_to_item(updated)


@router.get("/high-priority", response_model=List[ValidationItemResponse])
async def get_high_priority(db: aiosqlite.Connection = Depends(get_db)):
    rows = await (await db.execute(
        "SELECT * FROM validation_results WHERE priority='HIGH' ORDER BY domain, item_no"
    )).fetchall()
    return [_row_to_item(r) for r in rows]


@router.post("/seed")
async def manual_seed(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user),
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다")

    await db.execute("DELETE FROM validation_results")
    await db.commit()

    # Re-run seed (xlsx first, fallback to dummy)
    xlsx_path = "/app/docs/00. validation_plan.xlsx"
    items: List[tuple] = []
    source = "dummy"
    if os.path.exists(xlsx_path):
        try:
            items = _parse_xlsx(xlsx_path)
            if items:
                source = "xlsx"
        except Exception:
            items = []
    if not items:
        items = _generate_dummy_items()

    from core.db import get_db_path
    import aiosqlite as _aio
    async with _aio.connect(get_db_path()) as db2:
        db2.row_factory = _aio.Row
        await db2.executemany(
            "INSERT INTO validation_results (domain, item_no, area, item_name, priority, method, status) VALUES (?,?,?,?,?,?,?)",
            items,
        )
        await db2.commit()

    return {"message": f"재시드 완료: {len(items)}개 항목 (소스: {source})"}


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _row_to_item(row: aiosqlite.Row) -> ValidationItemResponse:
    keys = row.keys()
    return ValidationItemResponse(
        id=row["id"],
        domain=row["domain"],
        item_no=row["item_no"],
        item_name=row["item_name"],
        priority=row["priority"],
        status=row["status"],
        area=row["area"] if "area" in keys else None,
        method=row["method"] if "method" in keys else None,
        note=row["note"],
        assignee=row["assignee"],
        verified_at=row["verified_at"],
        verified_by=row["verified_by"],
    )
